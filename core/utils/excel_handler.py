"""
Excel utilities untuk import/export data
Digunakan di semua app untuk konsistensi
"""
import io
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.db import models
from django.core.exceptions import ValidationError


class ExcelExporter:
    """Export data Django model ke Excel file"""
    
    def __init__(self, model, queryset=None, columns=None, title=None):
        """
        Args:
            model: Django model class
            queryset: QuerySet untuk di-export (default: semua data)
            columns: List field names [(field_name, display_name), ...]
                    Jika None, ambil dari model fields
            title: Judul sheet (default: model name)
        """
        self.model = model
        self.queryset = queryset or model.objects.all()
        self.title = title or model._meta.verbose_name_plural
        self.columns = columns or self._get_default_columns()
    
    def _get_default_columns(self):
        """Get semua field dari model kecuali ManyToMany relations"""
        columns = []
        for field in self.model._meta.get_fields():
            # Skip many-to-many dan one-to-many relations
            if field.one_to_many or field.many_to_many:
                continue

            # Primary key dibuat otomatis oleh sistem, jadi tidak ikut export
            if getattr(field, "primary_key", False):
                continue
            
            # Untuk FK, gunakan field_id agar bisa di-import dengan ID
            if field.many_to_one:
                field_name = f"{field.name}_id"
                label = f"{getattr(field, 'verbose_name', field.name).title()} (ID)"
            else:
                field_name = field.name
                label = getattr(field, 'verbose_name', field.name).title()
            
            columns.append((field_name, label))
        return columns
    
    def export(self):
        """Export ke Excel dan return file bytes"""
        wb = Workbook()
        ws = wb.active
        ws.title = self.title[:31]  # Sheet name max 31 chars
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col_idx, (field_name, display_name) in enumerate(self.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = display_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        for row_idx, obj in enumerate(self.queryset, 2):
            for col_idx, (field_name, _) in enumerate(self.columns, 1):
                value = getattr(obj, field_name, None)
                
                # Format datetime
                if isinstance(value, datetime):
                    value = value.strftime('%d/%m/%Y %H:%M')
                elif isinstance(value, models.Model):
                    # Jangan sampai terjadi, karena sudah gunakan field_id
                    value = str(value)
                
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = data_alignment
                cell.border = border
        
        # Auto-adjust column widths
        for col_idx, (_, display_name) in enumerate(self.columns, 1):
            max_length = max(len(str(display_name)), 20)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
        
        # Return as bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


class ExcelImporter:
    """Import & validasi data dari Excel file"""
    
    def __init__(self, model, file_stream, columns=None, skip_empty_rows=True):
        """
        Args:
            model: Django model untuk import
            file_stream: File object atau bytes
            columns: List field names untuk mapping dengan Excel columns
            skip_empty_rows: Skip baris kosong (default: True)
        """
        self.model = model
        self.file_stream = file_stream
        self.columns = columns or self._get_default_columns()
        self.skip_empty_rows = skip_empty_rows
        self.errors = []
        self.preview_data = []
    
    def _get_default_columns(self):
        """Get semua field yang bisa di-import (kecuali many-to-many dan one-to-many)"""
        fields = []
        for field in self.model._meta.get_fields():
            # Skip many-to-many dan one-to-many relations
            if field.many_to_many or field.one_to_many:
                continue

            if getattr(field, "primary_key", False):
                continue
            
            # Untuk FK, gunakan field_id
            if field.many_to_one:
                fields.append(f"{field.name}_id")
            else:
                fields.append(field.name)
        return fields
    
    def _convert_field_id_to_field(self, field_name):
        """Convert field_id ke field name jika diperlukan"""
        if field_name.endswith('_id'):
            actual_field_name = field_name[:-3]  # Remove '_id'
            try:
                field = self.model._meta.get_field(actual_field_name)
                if field.many_to_one:
                    return actual_field_name, True  # field_name, is_fk=True
            except:
                pass
        return field_name, False
    
    def read_excel(self):
        """Parse Excel file dan return data"""
        try:
            if isinstance(self.file_stream, bytes):
                book = load_workbook(io.BytesIO(self.file_stream))
            else:
                book = load_workbook(self.file_stream)
            
            ws = book.active
            data = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                # Skip empty rows
                if self.skip_empty_rows and all(cell is None for cell in row):
                    continue
                
                row_data = {}
                for col_idx, field_name in enumerate(self.columns):
                    if col_idx < len(row):
                        row_data[field_name] = row[col_idx]
                    else:
                        row_data[field_name] = None
                
                data.append({
                    'row': row_idx,
                    'data': row_data
                })
            
            return data
        
        except Exception as e:
            self.errors.append(f"Error reading Excel: {str(e)}")
            return []
    
    def validate(self, data=None):
        """Validasi data sebelum import"""
        if data is None:
            data = self.read_excel()
        
        self.preview_data = []
        self.errors = []
        
        for item in data:
            row_num = item['row']
            row_data = item['data']
            errors = []
            
            try:
                # Process foreign keys - lookup by name or id
                processed_data = {}
                for field_col_name, value in row_data.items():
                    if value is None or value == '':
                        processed_data[field_col_name] = None
                        continue
                    
                    # Convert field_id ke field name jika FK
                    field_name, is_fk_field = self._convert_field_id_to_field(field_col_name)
                    field = self.model._meta.get_field(field_name)
                    
                    # Handle foreign keys (either field_id or direct field)
                    if is_fk_field or field.many_to_one:
                        related_model = field.related_model
                        # Try to find by id first, then by name
                        try:
                            # If value is numeric, try as id
                            if str(value).isdigit():
                                obj = related_model.objects.get(id=int(value))
                            else:
                                # Try to find by common name fields
                                name_fields = ['nama', 'name', 'title', 'pangkat', 'jabatan']
                                obj = None
                                for name_field in name_fields:
                                    if hasattr(related_model, name_field):
                                        try:
                                            obj = related_model.objects.get(**{name_field: str(value)})
                                            break
                                        except related_model.DoesNotExist:
                                            continue
                                if obj is None:
                                    # Try __str__ representation
                                    obj = related_model.objects.get(pk=value)
                            
                            # Store dengan field_name (bukan field_col_name)
                            processed_data[field_name] = obj
                        except (related_model.DoesNotExist, ValueError):
                            errors.append(f"Foreign key '{field_name}' dengan value '{value}' tidak ditemukan")
                            processed_data[field_name] = None
                    elif field.__class__.__name__ == 'DateField' and value:
                        # Parse date from Indonesian format DD/MM/YYYY
                        from datetime import datetime
                        try:
                            if isinstance(value, str):
                                # Try DD/MM/YYYY format
                                processed_data[field_name] = datetime.strptime(str(value), '%d/%m/%Y').date()
                            else:
                                processed_data[field_name] = value
                        except ValueError:
                            # If parsing fails, keep as is
                            processed_data[field_name] = value
                    else:
                        processed_data[field_name] = value
                
                # Create instance (tanpa save)
                instance = self.model(**processed_data)
                
                # Check if this will be an update (existing unique field)
                unique_fields = self._get_unique_fields()
                is_update = False
                if unique_fields:
                    for unique_field in unique_fields:
                        if unique_field in processed_data and processed_data[unique_field] is not None:
                            try:
                                existing = self.model.objects.get(**{unique_field: processed_data[unique_field]})
                                is_update = True
                                break
                            except self.model.DoesNotExist:
                                continue
                
                # Validate, but exclude unique fields if this is an update
                if is_update:
                    # For updates, skip unique validation by temporarily removing unique constraints
                    exclude_fields = [f for f in unique_fields if f in processed_data]
                    instance.full_clean(exclude=exclude_fields)
                else:
                    instance.full_clean()  # Django validation
                
                # Check jika ada FK errors
                if errors:
                    error_msg = '; '.join(errors)
                    self.preview_data.append({
                        'row': row_num,
                        'data': row_data,
                        'status': 'error',
                        'errors': errors
                    })
                    self.errors.append(f"Row {row_num}: {error_msg}")
                else:
                    self.preview_data.append({
                        'row': row_num,
                        'data': row_data,
                        'status': 'valid',
                        'errors': []
                    })
            
            except ValidationError as e:
                error_msg = '; '.join([f"{k}: {', '.join(v)}" for k, v in e.message_dict.items()])
                self.preview_data.append({
                    'row': row_num,
                    'data': row_data,
                    'status': 'error',
                    'errors': [error_msg]
                })
                self.errors.append(f"Row {row_num}: {error_msg}")
            
            except Exception as e:
                self.preview_data.append({
                    'row': row_num,
                    'data': row_data,
                    'status': 'error',
                    'errors': [str(e)]
                })
                self.errors.append(f"Row {row_num}: {str(e)}")
        
        return len(self.errors) == 0
    
    def _get_unique_fields(self):
        """Get fields that can be used for uniqueness check"""
        unique_fields = []
        for field in self.model._meta.get_fields():
            if field.primary_key or getattr(field, 'unique', False):
                unique_fields.append(field.name)
        return unique_fields
    
    def import_data(self, data=None):
        """Import data ke database (hanya yang valid)"""
        if data is None:
            data = self.read_excel()
        
        if not self.validate(data):
            return {
                'success': False,
                'imported': 0,
                'errors': self.errors,
                'preview': self.preview_data
            }
        
        imported = 0
        updated = 0
        unique_fields = self._get_unique_fields()
        
        try:
            for item in self.preview_data:
                if item['status'] == 'valid':
                    # Re-process data for import
                    row_data = item['data']
                    processed_data = {}
                    
                    for field_col_name, value in row_data.items():
                        if value is None or value == '':
                            processed_data[field_col_name] = None
                            continue
                        
                        # Convert field_id ke field name jika FK
                        field_name, is_fk_field = self._convert_field_id_to_field(field_col_name)
                        field = self.model._meta.get_field(field_name)
                        
                        # Handle foreign keys (either field_id or direct field)
                        if is_fk_field or field.many_to_one:
                            related_model = field.related_model
                            # Try to find by id first, then by name
                            try:
                                # If value is numeric, try as id
                                if str(value).isdigit():
                                    obj = related_model.objects.get(id=int(value))
                                else:
                                    # Try to find by common name fields
                                    name_fields = ['nama', 'name', 'title', 'pangkat', 'jabatan']
                                    obj = None
                                    for name_field in name_fields:
                                        if hasattr(related_model, name_field):
                                            try:
                                                obj = related_model.objects.get(**{name_field: str(value)})
                                                break
                                            except related_model.DoesNotExist:
                                                continue
                                    if obj is None:
                                        # Try __str__ representation
                                        obj = related_model.objects.get(pk=value)
                                
                                processed_data[field_name] = obj
                            except (related_model.DoesNotExist, ValueError):
                                raise ValueError(f"Foreign key '{field_name}' dengan value '{value}' tidak ditemukan")
                        elif field.__class__.__name__ == 'DateField' and value:
                            # Parse date from Indonesian format DD/MM/YYYY
                            from datetime import datetime
                            try:
                                if isinstance(value, str):
                                    # Try DD/MM/YYYY format
                                    processed_data[field_name] = datetime.strptime(str(value), '%d/%m/%Y').date()
                                else:
                                    processed_data[field_name] = value
                            except ValueError:
                                # If parsing fails, keep as is
                                processed_data[field_name] = value
                        else:
                            processed_data[field_name] = value
                    
                    # Use update_or_create with unique fields
                    if unique_fields:
                        # Use first unique field that is in the data (skip 'id')
                        unique_field = None
                        for uf in unique_fields:
                            if uf != 'id' and uf in processed_data and processed_data[uf] is not None:
                                unique_field = uf
                                break
                        
                        if unique_field:
                            obj, created = self.model.objects.update_or_create(
                                **{unique_field: processed_data[unique_field]},
                                defaults=processed_data
                            )
                            if created:
                                imported += 1
                            else:
                                updated += 1
                        else:
                            # If no suitable unique field, create new
                            self.model.objects.create(**processed_data)
                            imported += 1
                    else:
                        # No unique fields, create new
                        self.model.objects.create(**processed_data)
                        imported += 1
            
            return {
                'success': True,
                'imported': imported,
                'updated': updated,
                'errors': [],
                'preview': self.preview_data
            }
        
        except Exception as e:
            return {
                'success': False,
                'imported': imported,
                'updated': updated,
                'errors': [str(e)],
                'preview': self.preview_data
            }
