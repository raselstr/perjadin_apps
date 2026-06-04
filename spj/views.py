from decimal import Decimal
from datetime import datetime
from io import BytesIO

from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from types import SimpleNamespace
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.db.models import Count
from django.utils.html import format_html, format_html_join
from django.utils import timezone
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from core.crud.base import BaseCRUDView
from perintah.models import Pelaksana, PemberiTugas
from profiles.utils import get_active_opd_id

from .access import (
    filter_spj_queryset_for_user,
    is_spj_admin_user,
    is_spj_pengguna_user,
    is_spj_verifikator_user,
)
from .forms import (
    JenisSPJForm,
    LaporanPerjalananForm,
    PenginapanForm,
    PesawatForm,
    TransportForm,
    UangHarianForm,
    UangRepresentasiForm,
)
from .models import (
    JenisSPJ,
    LaporanPerjalanan,
    Penginapan,
    Pesawat,
    Transport,
    UangHarian,
    UangRepresentasi,
)
from .tables import (
    JenisSPJTable,
    LaporanPerjalananTable,
    PenginapanTable,
    PesawatTable,
    TransportTable,
    UangHarianTable,
    UangRepresentasiTable,
)


class SPJPelaksanaOptionsView(LoginRequiredMixin, View):
    @staticmethod
    def _param(request, name):
        value = request.GET.get(name)
        if value:
            return value
        suffix = f"-{name}"
        for key, item in request.GET.items():
            if key.endswith(suffix) and item:
                return item
        return None

    @staticmethod
    def _html_options(queryset, selected_id=None):
        options = [format_html('<option value="">---------</option>')]
        options.append(
            format_html_join(
                "",
                '<option value="{}"{}>{}</option>',
                (
                    (
                        item.id,
                        format_html(" selected") if str(item.id) == str(selected_id) else "",
                        f"SPT #{item.spt_id} - {item.nama}",
                    )
                    for item in queryset.distinct()
                ),
            )
        )
        return HttpResponse("".join(str(option) for option in options))

    def get(self, request):
        spt_id = self._param(request, "spt")
        model = self._param(request, "model")
        jenis_spj_id = self._param(request, "jenis_spj")
        instance_id = self._param(request, "instance")
        selected_id = self._param(request, "pelaksana")
        queryset = Pelaksana.objects.select_related("nama").order_by("nama__nama")

        if not spt_id:
            if request.headers.get("HX-Request"):
                return self._html_options(queryset.none())
            return JsonResponse({"results": []})

        queryset = queryset.filter(spt_id=spt_id)

        if is_spj_pengguna_user(request.user):
            queryset = queryset.filter(nama__nip=request.user.username)
        elif not is_spj_admin_user(request.user):
            active_opd_id = get_active_opd_id(request)
            if active_opd_id:
                queryset = queryset.filter(nama__opd_id=active_opd_id)

        used = None
        if model == "penginapan":
            used = Penginapan.objects.filter(spt_id=spt_id)
        elif model == "uang_harian":
            used = UangHarian.objects.filter(spt_id=spt_id)
        elif model == "uang_representasi":
            used = UangRepresentasi.objects.filter(spt_id=spt_id)
        elif model == "pesawat" and jenis_spj_id:
            used = Pesawat.objects.filter(
                spt_id=spt_id,
                jenis_spj_id=jenis_spj_id,
            )

        if used is not None:
            if instance_id:
                used = used.exclude(pk=instance_id)
            queryset = queryset.exclude(
                pk__in=used.values_list("pelaksana_id", flat=True)
            )

        if request.headers.get("HX-Request"):
            return self._html_options(queryset, selected_id)

        return JsonResponse({
            "results": [
                {"id": item.id, "text": f"SPT #{item.spt_id} - {item.nama}"}
                for item in queryset.distinct()
            ]
        })


class SPJCalculationView(LoginRequiredMixin, View):
    def get(self, request):
        jenis = request.GET.get("jenis")
        spt_id = request.GET.get("spt")
        pelaksana_id = request.GET.get("pelaksana")
        data = {"nilai": None, "total": None, "eligible": False}

        if not spt_id or not pelaksana_id:
            return JsonResponse(data)

        try:
            pelaksana = Pelaksana.objects.select_related(
                "spt",
                "spt__kota_tujuan",
                "spt__jenis_kegiatan",
                "nama",
                "nama__tingkat",
            ).get(pk=pelaksana_id, spt_id=spt_id)
        except Pelaksana.DoesNotExist:
            return JsonResponse(data)

        if jenis == "uang_harian":
            obj = UangHarian(spt=pelaksana.spt, pelaksana=pelaksana)
            nilai = obj.get_standar_maksimal()
            data["nilai"] = str(nilai) if nilai is not None else None
            data["total"] = (
                str(nilai * pelaksana.spt.lama_perjalanan)
                if nilai is not None else None
            )
            data["eligible"] = nilai is not None
        elif jenis == "representasi":
            obj = UangRepresentasi(spt=pelaksana.spt, pelaksana=pelaksana)
            nilai = obj.get_standar_maksimal()
            data["nilai"] = str(nilai) if nilai is not None else None
            data["eligible"] = nilai is not None

        return JsonResponse(data)


class SPJAvailableOptionsView(LoginRequiredMixin, View):
    def get(self, request):
        model = request.GET.get("model")
        if model != "transport":
            return JsonResponse({
                "unavailable_jenis_spj": [],
                "unavailable_jenis_transportasi": [],
            })

        spt_id = request.GET.get("spt")
        pelaksana_id = request.GET.get("pelaksana")
        jenis_spj_id = request.GET.get("jenis_spj")
        instance_id = request.GET.get("instance")

        queryset = filter_spj_queryset_for_user(
            Transport.objects.all(),
            request,
            "pelaksana__nama__nip",
        )
        if instance_id:
            queryset = queryset.exclude(pk=instance_id)

        unavailable_transport = []
        unavailable_jenis_spj = []

        if spt_id and pelaksana_id and jenis_spj_id:
            unavailable_transport = list(
                queryset.filter(
                    spt_id=spt_id,
                    pelaksana_id=pelaksana_id,
                    jenis_spj_id=jenis_spj_id,
                ).values_list("jenis_transportasi_id", flat=True)
            )

        if spt_id and pelaksana_id:
            total_transport_types = (
                Transport._meta.get_field("jenis_transportasi")
                .remote_field.model.objects.count()
            )
            if total_transport_types:
                used_by_jenis = (
                    queryset.filter(spt_id=spt_id, pelaksana_id=pelaksana_id)
                    .values("jenis_spj_id")
                    .annotate(
                        used_count=Count("jenis_transportasi_id", distinct=True)
                    )
                    .filter(used_count__gte=total_transport_types)
                )
                unavailable_jenis_spj = [
                    item["jenis_spj_id"] for item in used_by_jenis
                ]

        return JsonResponse({
            "unavailable_jenis_spj": unavailable_jenis_spj,
            "unavailable_jenis_transportasi": unavailable_transport,
        })


def _date_param(request, name):
    value = request.GET.get(name)
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _first_or_none(queryset):
    return queryset.first()


class SPJQuerysetMixin:
    template_name = "spj/page.html"
    template_form = "spj/form_page.html"
    template_delete = "spj/delete_page.html"

    def get_base_queryset(self):
        queryset = self.model.objects.select_related(
            "spt",
            "spt__kota_tujuan",
            "spt__jenis_kegiatan",
            "pelaksana",
            "pelaksana__nama",
            "pelaksana__nama__opd",
            "pelaksana__nama__tingkat",
        ).order_by("-id")
        return filter_spj_queryset_for_user(
            queryset,
            self.request,
            "pelaksana__nama__nip",
        ).distinct()

    def _is_locked_for_user(self, instance):
        return (
            getattr(instance, "verif_status", None) == "verified"
            and not is_spj_verifikator_user(self.request.user)
        )

    def form_view(self, request, pk=None):
        perm = self.get_permission()
        if pk:
            if not perm or not perm.can_edit:
                return self._forbidden(request)
        else:
            if not perm or not perm.can_add:
                return self._forbidden(request)

        instance = None
        if pk:
            instance = get_object_or_404(self.get_object_queryset(), pk=pk)
            if self._is_locked_for_user(instance):
                messages.error(
                    request,
                    "Berkas SPJ sudah diverifikasi dan tidak dapat diubah.",
                )
                if request.headers.get("HX-Request"):
                    return self._forbidden(request)
                return redirect(self.url_list)

        form = self.form_class(**self.get_form_kwargs(request, instance=instance))

        if request.method == "POST" and form.is_valid():
            action = "update" if instance else "add"
            obj = form.save(commit=False)
            if is_spj_verifikator_user(request.user):
                obj.mark_verification_user(request.user)
            obj.save()
            form.save_m2m()

            if request.headers.get("HX-Request"):
                return self._build_htmx_success_response(action)

            self._add_success_message(request, action)
            return redirect(self.url_list)

        context = {
            "form": form,
            "title": self.title,
            "permission": perm,
            "url_list": self.url_list,
            "is_multipart_form": form.is_multipart(),
        }

        if request.method == "POST" and request.headers.get("HX-Request"):
            return self._build_htmx_error_response(request, context, form)

        return render(request, self.template_form, context)

    def delete_view(self, request, pk):
        perm = self.get_permission()
        if not perm or not perm.can_delete:
            return self._forbidden(request)

        instance = get_object_or_404(self.get_object_queryset(), pk=pk)
        if self._is_locked_for_user(instance):
            messages.error(
                request,
                "Berkas SPJ sudah diverifikasi dan tidak dapat dihapus.",
            )
            if request.headers.get("HX-Request"):
                return self._forbidden(request)
            return redirect(self.url_list)

        if request.method == "POST":
            instance.delete()

            if request.headers.get("HX-Request"):
                return self._build_htmx_success_response("delete")

            self._add_success_message(request, "delete")
            return redirect(self.url_list)

        if request.headers.get("HX-Request"):
            return super().delete_view(request, pk)

        return render(request, self.template_delete, {
            "object": instance,
            "url_list": self.url_list,
            "title": self.title,
        })

class JenisSPJView(BaseCRUDView):
    model = JenisSPJ
    form_class = JenisSPJForm
    table_class = JenisSPJTable
    enable_excel = False

    title = "Jenis SPJ"
    url_list = "jenis_spj_list"
    url_action = "jenis_spj_action"
    url_action_pk = "jenis_spj_action_pk"


class PenginapanView(SPJQuerysetMixin, BaseCRUDView):
    model = Penginapan
    form_class = PenginapanForm
    table_class = PenginapanTable
    enable_excel = False

    title = "SPJ Penginapan"
    url_list = "penginapan_list"
    url_action = "penginapan_action"
    url_action_pk = "penginapan_action_pk"


class PesawatView(SPJQuerysetMixin, BaseCRUDView):
    model = Pesawat
    form_class = PesawatForm
    table_class = PesawatTable
    enable_excel = False

    title = "SPJ Pesawat"
    url_list = "pesawat_list"
    url_action = "pesawat_action"
    url_action_pk = "pesawat_action_pk"

    def get_base_queryset(self):
        return super().get_base_queryset().select_related(
            "jenis_spj",
            "lokasi_bandara",
            "lokasi_bandara__provinsi",
            "tujuan_bandara",
            "tujuan_bandara__provinsi",
        )


class UangHarianView(SPJQuerysetMixin, BaseCRUDView):
    model = UangHarian
    form_class = UangHarianForm
    table_class = UangHarianTable
    enable_excel = False

    title = "SPJ Uang Harian"
    url_list = "uang_harian_list"
    url_action = "uang_harian_action"
    url_action_pk = "uang_harian_action_pk"


class TransportView(SPJQuerysetMixin, BaseCRUDView):
    model = Transport
    form_class = TransportForm
    table_class = TransportTable
    enable_excel = False

    title = "SPJ Transport"
    url_list = "transport_list"
    url_action = "transport_action"
    url_action_pk = "transport_action_pk"

    def get_base_queryset(self):
        return super().get_base_queryset().select_related(
            "jenis_spj",
            "jenis_transportasi",
            "lokasi_berangkat",
            "tujuan",
        )


class UangRepresentasiView(SPJQuerysetMixin, BaseCRUDView):
    model = UangRepresentasi
    form_class = UangRepresentasiForm
    table_class = UangRepresentasiTable
    enable_excel = False

    title = "SPJ Uang Representasi"
    url_list = "uang_representasi_list"
    url_action = "uang_representasi_action"
    url_action_pk = "uang_representasi_action_pk"


class LaporanPerjalananView(SPJQuerysetMixin, BaseCRUDView):
    model = LaporanPerjalanan
    form_class = LaporanPerjalananForm
    table_class = LaporanPerjalananTable
    enable_excel = False

    title = "Laporan Perjalanan Dinas"
    url_list = "laporan_perjalanan_list"
    url_action = "laporan_perjalanan_action"
    url_action_pk = "laporan_perjalanan_action_pk"

    def get_base_queryset(self):
        queryset = self.model.objects.select_related(
            "spt",
            "spt__kota_tujuan",
            "spt__jenis_kegiatan",
            "pelaksana",
            "pelaksana__nama",
            "pelaksana__nama__opd",
            "pelaksana__nama__tingkat",
        ).order_by("-id")

        if is_spj_pengguna_user(self.request.user):
            return queryset.filter(
                spt__pelaksana__nama__nip=self.request.user.username,
            ).distinct()

        return filter_spj_queryset_for_user(
            queryset,
            self.request,
            "pelaksana__nama__nip",
        ).distinct()

    def get_permission(self):
        if is_spj_pengguna_user(self.request.user):
            return SimpleNamespace(
                can_view=True,
                can_add=True,
                can_edit=True,
                can_delete=False,
            )
        return super().get_permission()


class SPJReportView(LoginRequiredMixin, View):
    template_name = "spj/report.html"

    def _filter(self, queryset):
        return filter_spj_queryset_for_user(
            queryset,
            self.request,
            "pelaksana__nama__nip",
        ).distinct()

    def _build_rows(self, request):
        start_date = _date_param(request, "tgl1")
        end_date = _date_param(request, "tgl2")
        nama_pelaksana = (request.GET.get("nama") or "").strip()
        pelaksana = Pelaksana.objects.select_related(
            "spt",
            "spt__kota_tujuan",
            "spt__jenis_kegiatan",
            "nama",
            "nama__pangkat",
            "nama__tingkat",
            "nama__opd",
        ).order_by("spt__tgl_berangkat", "nama__nama")

        pelaksana = filter_spj_queryset_for_user(
            pelaksana,
            request,
            "nama__nip",
        ).distinct()

        if start_date:
            pelaksana = pelaksana.filter(spt__tgl_berangkat__gte=start_date)
        if end_date:
            pelaksana = pelaksana.filter(spt__tgl_berangkat__lte=end_date)
        if nama_pelaksana:
            pelaksana = pelaksana.filter(nama__nama__icontains=nama_pelaksana)

        rows = []
        for item in pelaksana:
            spt = item.spt
            pemberi = PemberiTugas.objects.filter(spt=spt).first()
            pesawat_berangkat = Pesawat.objects.filter(
                spt=spt,
                pelaksana=item,
                jenis_spj__jenis_spj__iexact="Berangkat",
            ).first()
            pesawat_kembali = Pesawat.objects.filter(
                spt=spt,
                pelaksana=item,
                jenis_spj__jenis_spj__iexact="Kembali",
            ).first()
            hotel = Penginapan.objects.filter(spt=spt, pelaksana=item).first()
            uang_harian = UangHarian.objects.filter(spt=spt, pelaksana=item).first()
            representasi = UangRepresentasi.objects.filter(spt=spt, pelaksana=item).first()
            transports = Transport.objects.filter(spt=spt, pelaksana=item).select_related("jenis_spj", "tujuan")
            transport_berangkat = transports.filter(jenis_spj__jenis_spj__iexact="Berangkat").first()
            transport_kembali = transports.filter(jenis_spj__jenis_spj__iexact="Kembali").first()

            total = sum([
                getattr(uang_harian, "total_biaya", Decimal("0")) or Decimal("0"),
                getattr(representasi, "total_biaya", Decimal("0")) or Decimal("0"),
                getattr(pesawat_berangkat, "total_biaya", Decimal("0")) or Decimal("0"),
                getattr(pesawat_kembali, "total_biaya", Decimal("0")) or Decimal("0"),
                getattr(hotel, "total_biaya", Decimal("0")) or Decimal("0"),
                getattr(transport_berangkat, "total_biaya", Decimal("0")) or Decimal("0"),
                getattr(transport_kembali, "total_biaya", Decimal("0")) or Decimal("0"),
            ], Decimal("0"))

            rows.append({
                "spt": spt,
                "pelaksana": item,
                "pemberi": pemberi,
                "pesawat_berangkat": pesawat_berangkat,
                "pesawat_kembali": pesawat_kembali,
                "hotel": hotel,
                "uang_harian": uang_harian,
                "representasi": representasi,
                "transport_berangkat": transport_berangkat,
                "transport_kembali": transport_kembali,
                "total": total,
            })

        return rows, start_date, end_date, nama_pelaksana

    def _row_values(self, idx, row):
        pegawai = row["pelaksana"].nama
        pangkat = pegawai.pangkat or "-"
        tingkat = pegawai.tingkat or "-"
        spt = row["spt"]
        pemberi = row["pemberi"]
        pb = row["pesawat_berangkat"]
        pk = row["pesawat_kembali"]
        hotel = row["hotel"]
        tb = row["transport_berangkat"]
        tk = row["transport_kembali"]
        return [
            idx,
            spt.tgl_berangkat.year if spt.tgl_berangkat else "",
            str(spt.jenis_kegiatan or ""),
            pegawai.nama,
            f"{pegawai.jabatan} / {pangkat} / {tingkat}",
            getattr(pemberi, "nomor_spt", "") or "-",
            getattr(pemberi, "tanggal_spt", "") or "",
            spt.tujuan_perjalanan_display or "",
            spt.lama_perjalanan,
            spt.tgl_berangkat,
            spt.tgl_kembali,
            getattr(pemberi, "nomor_spd", "") or "-",
            getattr(pemberi, "tanggal_spt", "") or "",
            getattr(row["uang_harian"], "total_biaya", 0) or 0,
            getattr(row["representasi"], "total_biaya", 0) or 0,
            getattr(pb, "nama_maskapai", "") or "-",
            getattr(pb, "nomor_tiket", "") or "-",
            getattr(pb, "kode_booking", "") or "-",
            getattr(pb, "tanggal_penerbangan", "") or "",
            getattr(pb, "harga_tiket", 0) or 0,
            getattr(pk, "nama_maskapai", "") or "-",
            getattr(pk, "nomor_tiket", "") or "-",
            getattr(pk, "kode_booking", "") or "-",
            getattr(pk, "tanggal_penerbangan", "") or "",
            getattr(pk, "harga_tiket", 0) or 0,
            f"{getattr(hotel, 'nama_hotel', '')} {getattr(hotel, 'alamat_hotel', '')}".strip() or "-",
            getattr(hotel, "tipe_kamar", "") or "-",
            getattr(hotel, "nomor_kamar", "") or "-",
            getattr(hotel, "tanggal_checkin", "") or "",
            getattr(hotel, "tanggal_checkout", "") or "",
            getattr(hotel, "lama_menginap", "") or "-",
            getattr(hotel, "harga_per_malam", 0) or 0,
            getattr(hotel, "total_biaya", 0) or 0,
            str(getattr(tb, "tujuan", "") or "-"),
            getattr(tb, "tanggal_berangkat", None) or spt.tgl_berangkat,
            getattr(tb, "biaya", 0) or 0,
            str(getattr(tk, "tujuan", "") or "-"),
            getattr(tk, "tanggal_berangkat", None) or spt.tgl_kembali,
            getattr(tk, "biaya", 0) or 0,
            row["total"],
        ]

    def _excel_response(self, rows):
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Pengawas"
        headers = [
            "NO", "TAHUN", "JENIS PERJALANAN", "NAMA",
            "JABATAN/GOL/TINGKAT BIAYA", "NO. SPT", "TANGGAL SPT",
            "TEMPAT TUJUAN", "JUMLAH HARI", "TGL BERANGKAT", "TGL KEMBALI",
            "NO. SPD", "TANGGAL SPD", "UANG HARIAN", "REPRESENTATIF",
            "MASKAPAI BERANGKAT", "NO TIKET", "KODE BOOKING",
            "TGL PENERBANGAN", "HARGA", "MASKAPAI KEMBALI", "NO TIKET",
            "KODE BOOKING", "TGL PENERBANGAN", "HARGA",
            "NAMA DAN LOKASI HOTEL", "TIPE KAMAR", "NOMOR KAMAR",
            "TGL CHECKIN", "TGL CHECKOUT", "LAMA", "HARGA PER MALAM",
            "TOTAL HOTEL", "TRANSPORT BERANGKAT", "HARGA",
            "TRANSPORT KEMBALI", "HARGA", "TOTAL",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for idx, row in enumerate(rows, start=1):
            ws.append([
                self._excel_value(value)
                for value in self._row_values(idx, row)
            ])
        stream = BytesIO()
        wb.save(stream)
        response = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="laporan-pengawas.xlsx"'
        return response

    def _pdf_response(self, rows):
        grand_total = sum((row["total"] for row in rows), Decimal("0"))
        response = render(self.request, "spj/report_print.html", {
            "title": "Laporan Pengawas",
            "rows": rows,
            "grand_total": grand_total,
            "auto_print": True,
        })
        response["X-Frame-Options"] = "SAMEORIGIN"
        return response

    @staticmethod
    def _excel_value(value):
        if hasattr(value, "strftime"):
            return value
        if isinstance(value, (int, float, Decimal)) or value is None:
            return value
        return str(value)

    def get(self, request):
        rows, start_date, end_date, nama_pelaksana = self._build_rows(request)
        export = request.GET.get("export")
        if export == "excel":
            return self._excel_response(rows)
        if export == "pdf":
            return self._pdf_response(rows)

        grand_total = sum((row["total"] for row in rows), Decimal("0"))
        return render(request, self.template_name, {
            "title": "Laporan Pengawas",
            "rows": rows,
            "grand_total": grand_total,
            "start_date": start_date,
            "end_date": end_date,
            "nama_pelaksana": nama_pelaksana,
        })


class LaporanPerjalananPrintView(LoginRequiredMixin, View):
    template_name = "spj/laporan_perjalanan_print.html"

    @staticmethod
    def _laporan_photos(laporan):
        photos = []
        for index in range(1, 5):
            image = getattr(laporan, f"foto_{index}", None)
            if image:
                photos.append({
                    "image": image,
                    "caption": f"Foto kegiatan {index}",
                })
        return photos

    def get_queryset(self, request):
        queryset = LaporanPerjalanan.objects.select_related(
            "spt",
            "spt__kota_tujuan",
            "spt__jenis_kegiatan",
            "pelaksana",
            "pelaksana__nama",
            "pelaksana__nama__opd",
        )
        if is_spj_pengguna_user(request.user):
            return queryset.filter(
                spt__pelaksana__nama__nip=request.user.username,
            ).distinct()
        return filter_spj_queryset_for_user(
            queryset,
            request,
            "pelaksana__nama__nip",
        ).distinct()

    def get(self, request, pk):
        laporan = get_object_or_404(self.get_queryset(request), pk=pk)
        context = {
            "laporan": laporan,
            "spt": laporan.spt,
            "pelaksana_list": laporan.spt.pelaksana.select_related(
                "nama",
                "nama__pangkat",
            ).all(),
            "laporan_photos": self._laporan_photos(laporan),
            "auto_print": request.GET.get("autoprint", "1") != "0",
        }
        response = render(request, self.template_name, context)
        response["X-Frame-Options"] = "SAMEORIGIN"
        response["Content-Security-Policy"] = (
            "default-src 'self'; "
            "script-src 'self' 'unsafe-inline'; "
            "style-src 'self' 'unsafe-inline'; "
            "img-src 'self' data: blob:; "
            "font-src 'self' data:; "
            "frame-ancestors 'self'; "
            "base-uri 'self'; "
            "form-action 'self'; "
            "object-src 'none'"
        )
        return response


class LaporanPerjalananPreviewView(LaporanPerjalananPrintView):
    template_name = "components/pdf/preview_modal.html"

    def get(self, request, pk):
        laporan = get_object_or_404(self.get_queryset(request), pk=pk)
        iframe_src = (
            reverse("laporan_perjalanan_print", args=[laporan.pk])
            + "?autoprint=0"
        )
        return render(request, self.template_name, {
            "title": "Preview Laporan Perjalanan",
            "document_code": "laporan",
            "preview_description": (
                "Tinjau Laporan Perjalanan terlebih dahulu sebelum dicetak."
            ),
            "iframe_src": iframe_src,
            "open_url": iframe_src,
            "frame_id": f"print-preview-frame-laporan-{laporan.pk}",
            "print_button_label": "Cetak Laporan",
        })
